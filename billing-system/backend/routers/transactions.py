import io
import os
import shutil
import re
import uuid
import random
import asyncio
from typing import List
from datetime import datetime
from decimal import Decimal
from bson import ObjectId
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, BackgroundTasks
from fastapi.responses import StreamingResponse
from database import get_db
from models import TransactionCreate
from auth import get_current_user
from utils import generate_invoice_pdf, send_whatsapp_invoice
from ai_utils import get_smart_product_match
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from zipfile import ZipFile
import traceback
from typing import Optional

def parse_date_string(date_str: Optional[str]) -> Optional[datetime]:
    if not date_str or str(date_str).strip() in ["", "-", "None"]:
        return None
    try:
        if re.match(r'^\d{4}-\d{2}-\d{2}', date_str):
            return datetime.strptime(date_str[:10], "%Y-%m-%d")
        parts = re.split(r'[/-]', date_str)
        if len(parts) == 3:
            d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
            if y < 100:
                y += 2000
            return datetime(y, m, d)
    except Exception:
        pass
    return None

def safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except Exception:
        return default

router = APIRouter()

UPLOAD_DIR = "uploads"
PAYMENT_PROOF_DIR = os.path.join(UPLOAD_DIR, "payment_proofs")
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(PAYMENT_PROOF_DIR, exist_ok=True)

@router.post("/manual")
async def create_transaction_manual(transaction: TransactionCreate, db=Depends(get_db), current_user=Depends(get_current_user)):
    tx_dict = transaction.model_dump()
    
    # If it's a single product manual entry, structure it for generate_invoice_pdf
    if not tx_dict.get("invoice_items"):
        # CASE-INSENSITIVE MATCH
        product = await db.products.find_one({"name": {"$regex": f"^{re.escape(tx_dict['product'])}$", "$options": "i"}})
        if not product:
            raise HTTPException(status_code=400, detail=f"Product '{tx_dict['product']}' (Case Insensitive) not found.")
        
        location = tx_dict.get("location", "India")
        price = Decimal(str(product["price_india"] if location == "India" else product["price_abroad"]))
        if price <= 0:
            raise HTTPException(status_code=400, detail=f"Product '{tx_dict['product']}' has 0 price. Blocked.")
        
        gst_rate = Decimal(str(product["gst_rate"] if location == "India" else 0))
        
        qty = Decimal("1")
        item_subtotal = price * qty
        gst_amount = item_subtotal * gst_rate / 100
        item_total = item_subtotal + gst_amount
        
        tx_dict["invoice_items"] = [{
            "name": product["name"],
            "qty": int(qty),
            "price": float(price),
            "gst_rate": float(gst_rate),
            "gst_amount": float(gst_amount),
            "total": float(item_total),
            "hsn": product.get("hsn_code", "9983")
        }]
        tx_dict["gst_breakdown"] = [{
            "rate": float(gst_rate),
            "cgst": float(gst_amount/2),
            "sgst": float(gst_amount/2),
            "total": float(gst_amount)
        }]
        tx_dict["amount"] = float(item_subtotal)
        tx_dict["gst_total"] = float(gst_amount)
        tx_dict["total_amount"] = float(item_total)

    # Calculate balance if paid_amount is present
    if tx_dict.get("paid_amount") is not None:
        tx_dict["balance"] = float(Decimal(str(tx_dict["total_amount"])) - Decimal(str(tx_dict["paid_amount"])))
    elif tx_dict.get("total_amount") is not None:
        # Default behavior: if not specified, assume unpaid or ask user? 
        # For manual entry, let's keep it None if not provided, or default to 0 balance if paid_amount = total_amount.
        # The USER screenshot shows some with null balance.
        pass
    
    tx_dict["status"] = "Pending"
    parsed_dt = parse_date_string(tx_dict.get("date"))
    tx_dict["timestamp"] = parsed_dt if parsed_dt else datetime.utcnow()
    tx_dict["added_by"] = current_user["username"]
    tx_dict["invoice_number"] = await get_next_sequence(db, "invoice_number")
    
    # Normalize and Update Customer
    normalized_name = tx_dict["name"].strip()
    await db.customers.update_one(
        {"phone": tx_dict["phone"], "name": {"$regex": f"^{re.escape(normalized_name)}$", "$options": "i"}},
        {"$set": {"name": normalized_name}, "$inc": {"total_spent": tx_dict["amount"], "total_transactions": 1}},
        upsert=True
    )
    
    result = await db.transactions.insert_one(tx_dict)
    tx_dict["id"] = str(result.inserted_id)
    
    # Generate Invoice immediately
    invoice_url = generate_invoice_pdf(tx_dict)
    await db.transactions.update_one({"_id": result.inserted_id}, {"$set": {"invoice_url": invoice_url}})
    tx_dict["invoice_url"] = invoice_url
    
    if "_id" in tx_dict:
        del tx_dict["_id"]
    return tx_dict

async def get_next_sequence(db, name):
    """Get next sequence value based on the highest existing number in the database"""
    if name == "invoice_number":
        highest_tx = await db.transactions.find_one({}, sort=[("invoice_number", -1)])
        if highest_tx and highest_tx.get("invoice_number"):
            try:
                return max(150614, int(highest_tx["invoice_number"]) + 1)
            except (ValueError, TypeError):
                pass
        return 150614
        
    counter = await db.counters.find_one_and_update(
        {"_id": name},
        {"$inc": {"sequence_value": 1}},
        upsert=True,
        return_document=True
    )
    return counter["sequence_value"]


@router.post("/upload")
async def upload_transactions(file: UploadFile = File(...), db=Depends(get_db), current_user=Depends(get_current_user)):
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="Invalid file type")
    
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        
        all_transactions = []
        batch_id = str(uuid.uuid4())
        
        # Calculate starting invoice number for this upload batch
        highest_tx = await db.transactions.find_one({}, sort=[("invoice_number", -1)])
        next_invoice_num = 150614
        if highest_tx and highest_tx.get("invoice_number"):
            try:
                next_invoice_num = max(150614, int(highest_tx["invoice_number"]) + 1)
            except (ValueError, TypeError):
                pass
        
        for sheet in wb.worksheets:
            rows = list(sheet.iter_rows(values_only=False))
            if not rows: continue
            
            # Check for New Format headers
            # 1. Normalize Headers (Strip spaces and lower case)
            header_vals = [str(c.value).lower().strip() if c.value else "" for c in rows[0]]
            
            # Robust mapping
            def find_idx(possible_names, default=-1):
                for name in possible_names:
                    if name.lower().strip() in header_vals:
                        return header_vals.index(name.lower().strip())
                return default

            col_map = {
                "date": find_idx(["date", "billing date", "time"], 0),
                "name": find_idx(["name", "customer", "customer name"], 1),
                "phone": find_idx(["phone", "contact", "contact no", "mobile"], 2),
                "transaction_id": find_idx(["txn id", "transaction id", "transaction_id", "ref id", "gpay id"], 3),
                "amount": find_idx(["amount", "total", "price", "paid", "paid amount", "received"], 4),
                "items": find_idx(["details", "product", "items"], 5),
                "location": find_idx(["location", "region"], -1),
                "shipping": find_idx(["shipping", "delivery"], -1)
            }

            is_new_format = col_map["location"] != -1 and col_map["items"] != -1
            
            # Detect start_idx (Skip headers)
            start_idx = 1
            for i, row in enumerate(rows):
                # If this row looks like a header (contains "name" or "phone" in actual value)
                row_vals = [str(c.value).lower() if c.value else "" for c in row]
                if any(x in row_vals for x in ["name", "phone", "contact", "customer", "txn id"]):
                    start_idx = i + 1
                    break

            current_date = "-" 
            
            # Cache product names for AI matching (OPTIMIZED: Outside row loop)
            cursor_p = db.products.find({}, {"name": 1})
            all_product_names = [p["name"] for p in await cursor_p.to_list(length=1000)]
            
            for row in rows[start_idx:]:
                if not any(c.value for c in row): continue 
                
                # 1. Capture Sticky Date
                d_val = row[col_map["date"]].value if col_map["date"] < len(row) else None
                if d_val:
                    if hasattr(d_val, "strftime"):
                        current_date = d_val.strftime('%d/%m/%y')
                    else:
                        d_str = str(d_val).strip()
                        # Match 01/01/2024 or 2024-01-01
                        if re.match(r'^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', d_str) or re.match(r'^\d{4}[/-]\d{1,2}[/-]\d{1,2}', d_str):
                            current_date = d_str
                        elif len(d_str) > 5: # Fallback for other date strings
                            current_date = d_str[:15]

                # 2. Extract Identity
                raw_name = str(row[col_map["name"]].value if len(row) > col_map["name"] else "").strip()
                raw_phone = str(row[col_map["phone"]].value if len(row) > col_map["phone"] else "").strip()
                raw_tx_id = str(row[col_map["transaction_id"]].value if len(row) > col_map["transaction_id"] else "").strip()
                raw_excel_amount = row[col_map["amount"]].value if col_map["amount"] != -1 and len(row) > col_map["amount"] else None
                
                if not raw_name or raw_name.lower() in ["none", "nan", "total", "customer name"]: continue
                
                phone = "".join(re.findall(r'\d+', raw_phone))
                if len(phone) > 12: phone = phone[:12]
                if not phone: continue

                # 3. Handle Product & Pricing Logic (USER Logic Implementation)
                items_str = ""
                location = "India"
                shipping = 0
                
                if is_new_format:
                    loc_raw = str(row[col_map["location"]].value or "India").strip().lower()
                    if loc_raw in ["india", "indina", "ind", "in"]:
                        location = "India"
                    else:
                        location = "Abroad"
                    items_str = str(row[col_map["items"]].value or "").strip()
                    shipping = float(row[col_map["shipping"]].value or 0) if col_map["shipping"] != -1 else 0
                else:
                    # Old Format fallback
                    items_str = str(row[col_map["items"]].value or "").strip()
                    location = "India"
                    shipping = 0

                # --- Calculation Engine (Decimal High Precision) ---
                parsed_items = []
                for part in items_str.split(","):
                    part = part.strip()
                    if not part: continue
                    match = re.match(r"(.*)\((\d+)\)", part)
                    if match:
                        name = match.group(1).strip()
                        qty = Decimal(match.group(2))
                        parsed_items.append((name, qty))
                    else:
                        parsed_items.append((part, Decimal("1")))

                invoice_items_details = []
                subtotal = Decimal("0")
                total_gst = Decimal("0")
                gst_summary = {} 


                for item_name, qty in parsed_items:
                    # 1. Try Exact/Case-Insensitive Match
                    clean_item_name = " ".join(item_name.split()).strip()
                    product = await db.products.find_one({"name": {"$regex": f"^{re.escape(clean_item_name)}$", "$options": "i"}})
                    
                    if not product:
                        # 2. SMART AI FALLBACK
                        suggested_name = await get_smart_product_match(clean_item_name, all_product_names)
                        if suggested_name:
                            product = await db.products.find_one({"name": suggested_name})
                    
                    if not product:
                        # Final attempt: search for the name inside the product name or vice versa (relaxed)
                        product = await db.products.find_one({"name": {"$regex": re.escape(clean_item_name), "$options": "i"}})
                    
                    if not product:
                        # Auto-create product to allow seamless excel upload
                        # Determine category based on product name keywords
                        cat = "Others"
                        lower_name = clean_item_name.lower()
                        if any(x in lower_name for x in ["reiki", "class", "course", "level", "learn"]):
                            cat = "Classes"
                        elif any(x in lower_name for x in ["tarot", "reading", "voice", "video", "appointment", "consult", "app", "pesi"]):
                            cat = "Tarot"
                        elif any(x in lower_name for x in ["healing", "session", "brahma", "fullmoon", "ammavasai"]):
                            cat = "Healing"
                        elif any(x in lower_name for x in ["medicine", "bach", "flower", "ml"]):
                            cat = "Medicine"
                        elif any(x in lower_name for x in ["bracelet", "pyramid", "stone", "crystal", "wand", "ball", "pendant", "malai", "tree", "hanging", "plate", "selenite", "pyrite", "quartz", "amethyst", "tourmaline", "aventurine", "citrine", "tiger", "carnelian", "lapis", "howlite"]):
                            cat = "Crystals"
                        elif "card" in lower_name:
                            cat = "Cards"
                        elif any(x in lower_name for x in ["knot", "ritual"]):
                            cat = "Rituals"

                        # Determine default GST rate based on category
                        g_rate = 18.0
                        if cat == "Crystals":
                            g_rate = 0.25
                        elif cat == "Medicine":
                            g_rate = 12.0
                        elif cat in ["Classes", "Healing"]:
                            g_rate = 5.0

                        # Determine HSN code based on category
                        h_code = "9983"
                        if cat == "Crystals":
                            h_code = "7117"
                        elif cat == "Medicine":
                            h_code = "3004"
                        elif cat == "Classes":
                            h_code = "9992"
                        elif cat == "Healing":
                            h_code = "9993"
                        elif cat == "Tarot":
                            h_code = "9983"
                        elif cat == "Cards":
                            h_code = "4901"

                        is_serv = cat in ["Tarot", "Classes", "Healing", "Rituals"]

                        # Calculate default base price based on excel row amount if available
                        price_in = 0.0
                        price_ab = 0.0
                        if raw_excel_amount is not None:
                            try:
                                amt_clean = re.sub(r'[^\d.]', '', str(raw_excel_amount))
                                if amt_clean:
                                    amt_dec = Decimal(amt_clean)
                                    items_count = Decimal(str(len(parsed_items)))
                                    unit_total = amt_dec / (items_count * qty)
                                    
                                    tax_mult = Decimal("1") + Decimal(str(g_rate)) / Decimal("100")
                                    price_in = float((unit_total / tax_mult).quantize(Decimal("0.01")))
                                    price_ab = float(unit_total.quantize(Decimal("0.01")))
                            except Exception:
                                pass

                        new_product = {
                            "name": clean_item_name,
                            "category": cat,
                            "price_india": price_in,
                            "price_abroad": price_ab,
                            "gst_rate": g_rate,
                            "hsn_code": h_code,
                            "is_service": is_serv
                        }
                        
                        insert_res = await db.products.insert_one(new_product)
                        new_product["_id"] = insert_res.inserted_id
                        product = new_product
                        # Keep cached list updated
                        all_product_names.append(clean_item_name)
                    
                    price = Decimal(str(product["price_india"] if location == "India" else product["price_abroad"]))
                    gst_rate = Decimal(str(product["gst_rate"] if location == "India" else 0))
                    
                    # Rounding each step
                    # PREFER EXCEL AMOUNT IF SINGLE ITEM (Legacy logic check)
                    # If the user specifically wants the Excel amount to override the DB total for single items:
                    # (Currently commented out to follow Turn 4 rule: ALWAYS USE OFFICIAL PRICE)
                    # if raw_excel_amount and len(parsed_items) == 1:
                    #     try:
                    #         item_total = Decimal(str(raw_excel_amount)).quantize(Decimal("0.01"))
                    #         item_subtotal = (item_total / (1 + gst_rate/100)).quantize(Decimal("0.01"))
                    #         gst_amount = item_total - item_subtotal
                    #         price = (item_subtotal / qty).quantize(Decimal("0.01"))
                    #     except:
                    #         item_subtotal = (price * qty).quantize(Decimal("0.01"))
                    #         gst_amount = (item_subtotal * gst_rate / 100).quantize(Decimal("0.01"))
                    # else:
                    
                    item_subtotal = (price * qty).quantize(Decimal("0.01"))
                    gst_amount = (item_subtotal * gst_rate / 100).quantize(Decimal("0.01"))
                    
                    item_total = item_subtotal + gst_amount
                    
                    subtotal += item_subtotal
                    total_gst += gst_amount
                    
                    # Grouping
                    gst_key = float(gst_rate)
                    gst_summary[gst_key] = gst_summary.get(gst_key, 0) + float(gst_amount)
                    
                    invoice_items_details.append({
                        "name": product["name"],
                        "qty": int(qty),
                        "price": float(price),
                        "gst_rate": float(gst_rate),
                        "gst_amount": float(gst_amount),
                        "total": float(item_total.quantize(Decimal("0.01"))),
                        "hsn": product.get("hsn_code", product.get("hsn", "9983"))
                    })

                grand_total = subtotal + total_gst + Decimal(str(shipping)).quantize(Decimal("0.01"))
                
                gst_breakdown = []
                for rate, val in gst_summary.items():
                    gst_breakdown.append({
                        "rate": rate,
                        "cgst": val / 2,
                        "sgst": val / 2,
                        "total": val
                    })

                # Clean Transaction ID
                first_id = raw_tx_id.split("\n")[0].split("\r")[0].strip()
                clean_tx_id = "".join([c if c.isalnum() else "_" for c in first_id]).upper()[:30]
                if not clean_tx_id or clean_tx_id.lower() in ["nan", "total"]: continue

                # Check if this exact transaction (ID + Customer) already exists to avoid duplicates
                existing = await db.transactions.find_one({
                    "transaction_id": clean_tx_id, 
                    "name": raw_name
                })
                if existing: continue

                # Get Sequential Invoice Number
                invoice_num = next_invoice_num
                next_invoice_num += 1

                # Calculate Balance (Official Total - Excel Paid Amount)
                paid_val = None
                balance = None
                if raw_excel_amount is not None and str(raw_excel_amount).strip() != "":
                    try:
                        clean_amt_str = re.sub(r'[^\d.]', '', str(raw_excel_amount))
                        if clean_amt_str:
                            paid_val_dec = Decimal(clean_amt_str).quantize(Decimal("0.01"))
                            paid_val = float(paid_val_dec)
                            bal = float(grand_total - paid_val_dec)
                            balance = bal
                    except Exception:
                        pass

                # 4. Build Record
                tx_obj = {
                    "name": raw_name,
                    "phone": phone,
                    "transaction_id": clean_tx_id,
                    "amount": float(subtotal), 
                    "product": items_str, 
                    "date": current_date, 
                    "location": location,
                    "invoice_items": invoice_items_details,
                    "gst_breakdown": gst_breakdown,
                    "shipping": float(shipping),
                    "gst_total": float(total_gst),
                    "total_amount": float(grand_total),
                    "paid_amount": paid_val,
                    "balance": balance,
                    "status": "Verified",
                    "timestamp": parse_date_string(current_date) or datetime.utcnow(),
                    "invoice_number": invoice_num,
                    "added_by": current_user["username"],
                    "batch_id": batch_id
                }
                all_transactions.append(tx_obj)

        if all_transactions:
            result = await db.transactions.insert_many(all_transactions)
            for i, tx in enumerate(all_transactions):
                tx["_id"] = result.inserted_ids[i]
                # Normalize and Update Customer
                normalized_name = tx["name"].strip()
                await db.customers.update_one(
                    {"phone": tx["phone"], "name": {"$regex": f"^{re.escape(normalized_name)}$", "$options": "i"}},
                    {"$set": {"name": normalized_name}, "$inc": {"total_spent": tx["amount"], "total_transactions": 1}},
                    upsert=True
                )
                invoice_url = generate_invoice_pdf(tx)
                await db.transactions.update_one({"_id": tx["_id"]}, {"$set": {"invoice_url": invoice_url}})

        return {"message": f"Processed {len(all_transactions)} transactions", "batch_id": batch_id}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/")
async def get_transactions(skip: int = 0, limit: int = 50, status: str = None, search: str = None, latest_batch_only: bool = False, year: str = None, sort_by: str = "date", sort_order: str = "desc", db=Depends(get_db), current_user=Depends(get_current_user)):
    query = {}
    if status and status != "All": query["status"] = status
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
            {"transaction_id": {"$regex": search, "$options": "i"}},
            {"product": {"$regex": search, "$options": "i"}}
        ]
    if latest_batch_only:
        latest = await db.transactions.find_one(sort=[("timestamp", -1)])
        if latest: query["batch_id"] = latest.get("batch_id")
        
    if year and year != "All":
        if "-" in year:
            try:
                parts = year.split("-")
                start_yy = int(parts[0])
                end_yy = int(parts[1])
                # Convert 2-digit years to 4-digit years (assuming 2000s)
                start_year = 2000 + start_yy
                end_year = 2000 + end_yy
                # Financial year starts April 1st of start_year
                start_date = datetime(start_year, 4, 1)
                # Financial year ends March 31st of end_year, so exclusive boundary is April 1st of end_year
                end_date = datetime(end_year, 4, 1)
                query["timestamp"] = {"$gte": start_date, "$lt": end_date}
            except: pass
        else:
            try:
                year_int = int(year)
                start_date = datetime(year_int, 1, 1)
                end_date = datetime(year_int + 1, 1, 1)
                query["timestamp"] = {"$gte": start_date, "$lt": end_date}
            except: pass
            
    # Filter out future transactions by default (cap at current date/time + 1 day for timezone safety)
    from datetime import timedelta
    now_limit = datetime.utcnow() + timedelta(days=1)
    if "timestamp" in query:
        if isinstance(query["timestamp"], dict):
            upper_bound = query["timestamp"].get("$lt") or query["timestamp"].get("$lte")
            if upper_bound and upper_bound > now_limit:
                query["timestamp"]["$lt"] = now_limit
            elif not upper_bound:
                query["timestamp"]["$lte"] = now_limit
    else:
        query["timestamp"] = {"$lte": now_limit}
            
    # Determine sorting field and direction
    sort_field = "timestamp"
    if sort_by == "amount":
        sort_field = "total_amount"
    elif sort_by == "bill_no":
        sort_field = "invoice_number"
    elif sort_by == "customer":
        sort_field = "name"
        
    direction = -1 if sort_order == "desc" else 1
    cursor = db.transactions.find(query).sort([(sort_field, direction)]).skip(skip).limit(limit)
    transactions = await cursor.to_list(length=limit)
    
    serialized = []
    for tx in transactions:
        tx["id"] = str(tx["_id"])
        del tx["_id"]
        if "timestamp" in tx and hasattr(tx["timestamp"], "isoformat"):
            tx["timestamp"] = tx["timestamp"].isoformat()
        serialized.append(tx)
        
    total = await db.transactions.count_documents(query)
    return {"total": total, "items": serialized}

@router.delete("/{tx_id}")
async def delete_transaction(tx_id: str, db=Depends(get_db), current_user=Depends(get_current_user)):
    tx = await db.transactions.find_one({"_id": ObjectId(tx_id)})
    if tx:
        await db.customers.update_one(
            {"phone": tx["phone"]},
            {"$inc": {"total_spent": -tx["amount"], "total_transactions": -1}}
        )
        await db.customers.delete_many({"total_transactions": {"$lte": 0}})
        await db.transactions.delete_one({"_id": ObjectId(tx_id)})
    return {"message": "Deleted"}

@router.put("/{tx_id}")
async def update_transaction(tx_id: str, payload: dict, db=Depends(get_db), current_user=Depends(get_current_user)):
    # Find existing transaction
    tx = await db.transactions.find_one({"_id": ObjectId(tx_id)})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
        
    old_phone = tx.get("phone")
    old_name = tx.get("name", "").strip()
    old_amount = tx.get("amount", 0)
    old_invoice_url = tx.get("invoice_url")
    
    # Update transaction details, excluding read-only metadata fields
    exclude_fields = ["_id", "id", "added_by", "timestamp", "invoice_url", "payment_proof_url", "payment_proof_filename", "payment_proof_uploaded_at", "invoice_number"]
    update_data = {k: v for k, v in payload.items() if k not in exclude_fields}
    
    # Recalculate balance if total_amount or paid_amount is updated
    if "total_amount" in update_data or "paid_amount" in update_data:
        total = update_data.get("total_amount", tx.get("total_amount", 0))
        paid = update_data.get("paid_amount", tx.get("paid_amount"))
        if paid is not None:
            update_data["balance"] = float(Decimal(str(total)) - Decimal(str(paid)))
            
        # Scale pricing, base amount, and items proportionally if total_amount is changed
        old_total = tx.get("total_amount", 0)
        if float(total) != float(old_total):
            # Resolve GST rate
            gst_rate = tx.get("gst_rate")
            if gst_rate is None:
                if tx.get("invoice_items"):
                    gst_rate = tx["invoice_items"][0].get("gst_rate", 18.0)
                elif tx.get("gst_breakdown"):
                    gst_rate = tx["gst_breakdown"][0].get("rate", 18.0)
                else:
                    gst_rate = 18.0
            
            total_dec = Decimal(str(total))
            gst_rate_dec = Decimal(str(gst_rate))
            base_amount = total_dec / (Decimal("1") + (gst_rate_dec / Decimal("100")))
            gst_total = total_dec - base_amount
            
            update_data["amount"] = float(base_amount)
            update_data["gst_total"] = float(gst_total)
            
            # If there are structured invoice items, scale them proportionally
            if tx.get("invoice_items") and float(old_total) > 0:
                ratio = float(total) / float(old_total)
                new_items = []
                breakdown_map = {}
                
                for item in tx.get("invoice_items", []):
                    item_total = Decimal(str(item.get("total", 0))) * Decimal(str(ratio))
                    item_gst_rate = Decimal(str(item.get("gst_rate", 0)))
                    qty = Decimal(str(item.get("qty", 1)))
                    price_each = item_total / (Decimal("1") + (item_gst_rate / Decimal("100"))) / qty
                    item_subtotal = price_each * qty
                    item_gst = item_total - item_subtotal
                    
                    new_items.append({
                        "name": item["name"],
                        "qty": int(qty),
                        "price": float(price_each),
                        "gst_rate": float(item_gst_rate),
                        "gst_amount": float(item_gst),
                        "total": float(item_total),
                        "hsn": item.get("hsn", "9983")
                    })
                    
                    rate_val = float(item_gst_rate)
                    if rate_val not in breakdown_map:
                        breakdown_map[rate_val] = Decimal("0")
                    breakdown_map[rate_val] += item_gst
                    
                update_data["invoice_items"] = new_items
                
                new_breakdown = []
                for rate_val, gst_val in breakdown_map.items():
                    new_breakdown.append({
                        "rate": rate_val,
                        "cgst": float(gst_val / 2),
                        "sgst": float(gst_val / 2),
                        "total": float(gst_val)
                    })
                update_data["gst_breakdown"] = new_breakdown

    result = await db.transactions.update_one({"_id": ObjectId(tx_id)}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Transaction not found")
        
    # Sync customers collection if identifiers or transaction subtotal amount changed
    new_phone = update_data.get("phone", old_phone)
    new_name = update_data.get("name", old_name).strip()
    new_amount = update_data.get("amount", old_amount)
    
    if old_phone != new_phone or old_name.lower() != new_name.lower() or old_amount != new_amount:
        if old_phone and old_name:
            await db.customers.update_one(
                {"phone": old_phone, "name": {"$regex": f"^{re.escape(old_name)}$", "$options": "i"}},
                {"$inc": {"total_spent": -old_amount, "total_transactions": -1}}
            )
        if new_phone and new_name:
            await db.customers.update_one(
                {"phone": new_phone, "name": {"$regex": f"^{re.escape(new_name)}$", "$options": "i"}},
                {"$set": {"name": new_name}, "$inc": {"total_spent": new_amount, "total_transactions": 1}},
                upsert=True
            )
        await db.customers.delete_many({"total_transactions": {"$lte": 0}})
        
    # Regenerate invoice PDF and update URL
    updated_tx = await db.transactions.find_one({"_id": ObjectId(tx_id)})
    if updated_tx:
        try:
            new_invoice_url = generate_invoice_pdf(updated_tx)
            await db.transactions.update_one({"_id": ObjectId(tx_id)}, {"$set": {"invoice_url": new_invoice_url}})
            
            # Safely remove the old PDF file if it is different
            if old_invoice_url and old_invoice_url != new_invoice_url:
                old_pdf_path = os.path.join(os.getcwd(), old_invoice_url.lstrip("/"))
                if os.path.exists(old_pdf_path):
                    try:
                        os.remove(old_pdf_path)
                    except Exception as e:
                        print(f"Error removing old PDF: {e}")
        except Exception as e:
            print(f"Error regenerating invoice: {e}")
            
    return {"message": "Updated"}

@router.post("/bulk-delete")
async def bulk_delete(payload: dict, db=Depends(get_db), current_user=Depends(get_current_user)):
    is_delete_all = payload.get("deleteAll", False)
    
    if is_delete_all:
        await db.transactions.delete_many({})
        await db.customers.delete_many({})
        await db.whatsapp_batches.delete_many({})
        return {"message": "Database wiped clean."}
    else:
        ids_raw = payload.get("ids", [])
        if not ids_raw:
            return {"message": "No IDs provided"}
            
        ids = [ObjectId(i) for i in ids_raw]
        cursor = db.transactions.find({"_id": {"$in": ids}})
        transactions_to_delete = await cursor.to_list(length=len(ids))
        
        for tx in transactions_to_delete:
            phone = tx.get("phone")
            amount = tx.get("amount", 0)
            if phone:
                await db.customers.update_one(
                    {"phone": phone},
                    {"$inc": {"total_spent": -amount, "total_transactions": -1}}
                )
        
        await db.customers.delete_many({"total_transactions": {"$lte": 0}})
        await db.transactions.delete_many({"_id": {"$in": ids}})
        return {"message": f"Deleted {len(ids)} transactions."}

@router.post("/{tx_id}/generate-invoice")
async def create_invoice(tx_id: str, db=Depends(get_db)):
    tx = await db.transactions.find_one({"_id": ObjectId(tx_id)})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
        
    try:
        url = generate_invoice_pdf(tx)
        await db.transactions.update_one({"_id": ObjectId(tx_id)}, {"$set": {"invoice_url": url}})
        return {"url": url}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/{tx_id}/payment-proof")
async def upload_payment_proof(tx_id: str, file: UploadFile = File(...), db=Depends(get_db), current_user=Depends(get_current_user)):
    """Upload an image or PDF as payment proof for a transaction."""
    tx = await db.transactions.find_one({"_id": ObjectId(tx_id)})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    # Validate file type
    allowed_types = [".jpg", ".jpeg", ".png", ".gif", ".webp", ".pdf"]
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in allowed_types:
        raise HTTPException(status_code=400, detail=f"File type '{ext}' not allowed. Use: {', '.join(allowed_types)}")
    
    # Delete old proof file if exists
    old_proof = tx.get("payment_proof_url")
    if old_proof:
        old_path = os.path.join(os.getcwd(), old_proof.lstrip("/"))
        if os.path.exists(old_path):
            os.remove(old_path)
    
    # Save file with unique name
    safe_filename = f"{tx_id}_{uuid.uuid4().hex[:8]}{ext}"
    file_path = os.path.join(PAYMENT_PROOF_DIR, safe_filename)
    
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)
    
    # Store relative URL path
    proof_url = f"/uploads/payment_proofs/{safe_filename}"
    
    await db.transactions.update_one(
        {"_id": ObjectId(tx_id)},
        {"$set": {
            "payment_proof_url": proof_url,
            "payment_proof_filename": file.filename,
            "payment_proof_uploaded_at": datetime.utcnow()
        }}
    )
    
    return {"message": "Payment proof uploaded", "url": proof_url, "filename": file.filename}

@router.delete("/{tx_id}/payment-proof")
async def delete_payment_proof(tx_id: str, db=Depends(get_db), current_user=Depends(get_current_user)):
    """Delete payment proof for a transaction."""
    tx = await db.transactions.find_one({"_id": ObjectId(tx_id)})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    proof_url = tx.get("payment_proof_url")
    if proof_url:
        file_path = os.path.join(os.getcwd(), proof_url.lstrip("/"))
        if os.path.exists(file_path):
            os.remove(file_path)
    
    await db.transactions.update_one(
        {"_id": ObjectId(tx_id)},
        {"$unset": {"payment_proof_url": "", "payment_proof_filename": "", "payment_proof_uploaded_at": ""}}
    )
    
    return {"message": "Payment proof removed"}

@router.post("/{tx_id}/send-whatsapp")
async def send_whatsapp(tx_id: str, db=Depends(get_db)):
    tx = await db.transactions.find_one({"_id": ObjectId(tx_id)})
    res = send_whatsapp_invoice(tx["phone"], tx["invoice_url"])
    if res.get("status") != "success":
        raise HTTPException(status_code=500, detail=res.get("message", "Failed to send WhatsApp message"))
    return res

async def process_bulk_whatsapp(batch_id: str, ids: List[str], db):
    await db.whatsapp_batches.update_one(
        {"batch_id": batch_id},
        {"$set": {"status": "processing", "total": len(ids), "processed": 0}}
    )

    for i, tx_id in enumerate(ids):
        try:
            tx = await db.transactions.find_one({"_id": ObjectId(tx_id)})
            if not tx: continue
            
            await db.whatsapp_batches.update_one(
                {"batch_id": batch_id, "items.tx_id": tx_id},
                {"$set": {"items.$.status": "sending"}}
            )

            invoice_url = tx.get("invoice_url")
            if not invoice_url:
                invoice_url = generate_invoice_pdf(tx)
                await db.transactions.update_one({"_id": ObjectId(tx_id)}, {"$set": {"invoice_url": invoice_url}})
            
            res = send_whatsapp_invoice(tx["phone"], invoice_url)
            
            if res.get("status") == "success":
                await db.whatsapp_batches.update_one(
                    {"batch_id": batch_id, "items.tx_id": tx_id},
                    {"$set": {"items.$.status": "sent", "items.$.sent_at": datetime.utcnow()}}
                )
            else:
                await db.whatsapp_batches.update_one(
                    {"batch_id": batch_id, "items.tx_id": tx_id},
                    {"$set": {"items.$.status": "error", "items.$.error": res.get("message", "Failed to send")}}
                )
            await db.whatsapp_batches.update_one({"batch_id": batch_id}, {"$inc": {"processed": 1}})
            
            if i < len(ids) - 1:
                await asyncio.sleep(random.uniform(30, 50))
                
        except Exception as e:
            await db.whatsapp_batches.update_one(
                {"batch_id": batch_id, "items.tx_id": tx_id},
                {"$set": {"items.$.status": "error", "items.$.error": str(e)}}
            )

    await db.whatsapp_batches.update_one({"batch_id": batch_id}, {"$set": {"status": "completed", "completed_at": datetime.utcnow()}})

@router.post("/bulk-whatsapp")
async def bulk_send_whatsapp(payload: dict, background_tasks: BackgroundTasks, db=Depends(get_db)):
    ids = payload.get("ids", [])
    batch_id = str(uuid.uuid4())
    
    batch_items = []
    for tx_id in ids:
        tx = await db.transactions.find_one({"_id": ObjectId(tx_id)})
        if tx:
            batch_items.append({
                "tx_id": tx_id,
                "name": tx["name"],
                "phone": tx["phone"],
                "status": "queued"
            })
            
    await db.whatsapp_batches.insert_one({
        "batch_id": batch_id,
        "status": "queued",
        "total": len(ids),
        "processed": 0,
        "items": batch_items,
        "created_at": datetime.utcnow()
    })

    background_tasks.add_task(process_bulk_whatsapp, batch_id, ids, db)
    return {"message": f"Bulk sending started", "batch_id": batch_id}

@router.get("/whatsapp/batch/{batch_id}")
async def get_whatsapp_batch(batch_id: str, db=Depends(get_db)):
    batch = await db.whatsapp_batches.find_one({"batch_id": batch_id})
    if not batch: raise HTTPException(status_code=404, detail="Batch not found")
    batch["id"] = str(batch["_id"])
    del batch["_id"]
    return batch

@router.get("/batch-status")
async def get_current_batch_status(db=Depends(get_db)):
    batch = await db.whatsapp_batches.find_one(sort=[("created_at", -1)])
    if not batch: return {"status": "none", "items": []}
    batch["id"] = str(batch["_id"])
    del batch["_id"]
    return batch

@router.post("/bulk-export")
async def bulk_export(payload: dict, db=Depends(get_db)):
    ids = [ObjectId(i) for i in payload.get("ids", [])]
    cursor = db.transactions.find({"_id": {"$in": ids}})
    txs = await cursor.to_list(length=len(ids))
    
    zip_buffer = io.BytesIO()
    with ZipFile(zip_buffer, "w") as zip_file:
        for tx in txs:
            invoice_path = tx.get("invoice_url") or generate_invoice_pdf(tx)
            file_path = os.path.join(os.getcwd(), invoice_path.lstrip("/"))
            if os.path.exists(file_path):
                safe_name = f"Invoice_{tx.get('transaction_id', str(tx['_id']))}.pdf"
                zip_file.write(file_path, safe_name)
    
    zip_buffer.seek(0)
    return StreamingResponse(zip_buffer, media_type="application/x-zip-compressed", headers={"Content-Disposition": f"attachment; filename=Invoices.zip"})

@router.post("/export-analytics")
async def export_analytics(payload: dict, db=Depends(get_db), current_user=Depends(get_current_user)):
    """Export analytics report as Excel matching the template format with two sheets: 'Sale Report' and 'Sale Items'."""
    from openpyxl.utils import get_column_letter
    from fastapi.responses import JSONResponse
    
    start_date = payload.get("start_date")  # "YYYY-MM-DD"
    end_date = payload.get("end_date")      # "YYYY-MM-DD"
    products = payload.get("products", [])   # list of product name strings
    
    # Defensive backend date validation
    if start_date and end_date:
        if start_date > end_date:
            return JSONResponse(
                status_code=400,
                content={
                    "success": False,
                    "message": "End date cannot be earlier than start date."
                }
            )
        
    query = {}
    
    # Date filter — parse date field (stored as DD/MM/YY string) via timestamp
    if start_date or end_date:
        date_filter = {}
        if start_date:
            try:
                date_filter["$gte"] = datetime.strptime(start_date, "%Y-%m-%d")
            except ValueError:
                pass
        if end_date:
            try:
                # End of day
                end_dt = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
                date_filter["$lte"] = end_dt
            except ValueError:
                pass
        if date_filter:
            query["timestamp"] = date_filter
    
    # Product filter — match any product in the items string
    if products and len(products) > 0:
        product_patterns = [{"product": {"$regex": re.escape(p), "$options": "i"}} for p in products]
        if "$or" in query:
            existing_or = query.pop("$or")
            query["$and"] = [{"$or": existing_or}, {"$or": product_patterns}]
        else:
            query["$or"] = product_patterns
    
    cursor = db.transactions.find(query).sort([("timestamp", -1)])
    transactions = await cursor.to_list(length=100000)
    
    if not transactions or len(transactions) == 0:
        return JSONResponse(
            status_code=404,
            content={
                "success": False,
                "message": "No transactions found for the selected filters."
            }
        )
    
    # Fetch all products to determine if an item is a service for Sheet 2 Unit column
    products_cursor = db.products.find({}, {"name": 1, "is_service": 1})
    products_list = await products_cursor.to_list(length=1000)
    service_map = {p["name"].lower().strip(): p.get("is_service", False) for p in products_list}
    
    # Create Excel workbook
    wb = Workbook()
    
    # Helper to format transaction date
    def format_tx_date(tx):
        date_str = tx.get("date")
        if date_str:
            try:
                if re.match(r'^\d{4}-\d{2}-\d{2}', date_str):
                    dt = datetime.strptime(date_str[:10], "%Y-%m-%d")
                    return dt.strftime("%d/%m/%Y")
                parts = re.split(r'[/-]', date_str)
                if len(parts) == 3:
                    d, m, y = parts[0], parts[1], parts[2]
                    if len(y) == 2:
                        y = "20" + y
                    return f"{int(d):02d}/{int(m):02d}/{y}"
            except:
                pass
        ts = tx.get("timestamp")
        if ts:
            if hasattr(ts, "strftime"):
                return ts.strftime("%d/%m/%Y")
        return datetime.utcnow().strftime("%d/%m/%Y")

    # Helper to format invoice number
    def get_invoice_number_formatted(tx):
        inv_num = tx.get("invoice_number")
        if not inv_num:
            return ""
        
        date_obj = tx.get("timestamp")
        if not date_obj:
            date_str = tx.get("date")
            if date_str:
                try:
                    parts = re.split(r'[/-]', date_str)
                    if len(parts) == 3:
                        d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
                        if y < 100:
                            y += 2000
                        date_obj = datetime(y, m, d)
                except:
                    pass
        if not date_obj:
            date_obj = datetime.utcnow()
            
        year = date_obj.year
        month = date_obj.month
        if month >= 4:
            fy = f"{year % 100:02d}-{(year + 1) % 100:02d}"
        else:
            fy = f"{(year - 1) % 100:02d}-{year % 100:02d}"
            
        return f"ZH-FY{fy}/{inv_num}"

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="e2e8f0"),
        right=Side(style="thin", color="e2e8f0"),
        top=Side(style="thin", color="e2e8f0"),
        bottom=Side(style="thin", color="e2e8f0")
    )
    data_font = Font(name="Calibri", size=10)
    total_font = Font(name="Calibri", bold=True, size=11, color="10b981")
    total_fill = PatternFill(start_color="f0fdf4", end_color="f0fdf4", fill_type="solid")
    
    # ----------------------------------------------------
    # SHEET 1: Sale Report (Invoice summary ledger)
    # ----------------------------------------------------
    ws1 = wb.active
    ws1.title = "Sale Report"
    
    # Row 0 (openpyxl Row 1): Generation timestamp
    gen_time_str = datetime.now().strftime('%b %d,%Y at %I:%M %p').replace('AM', 'am').replace('PM', 'pm')
    ws1.cell(row=1, column=1, value=f"Generated on {gen_time_str}").font = Font(name="Calibri", size=10, italic=True, color="64748b")
    
    # Row 1 (openpyxl Row 2): Empty spacer row
    
    # Row 2 (openpyxl Row 3): UserName metadata
    ws1.cell(row=3, column=1, value="UserName").font = Font(name="Calibri", size=10, bold=True)
    ws1.cell(row=3, column=2, value=current_user["username"]).font = Font(name="Calibri", size=10)
    
    # Row 3 (openpyxl Row 4): Column Headers
    headers1 = ['Date', 'Party Name', 'Phone No.', "Party's GSTIN No.", 'Order No.', 'Invoice No.', 'Transaction Type', 'Total Amount', 'Payment Type', 'Received Amount', 'Balance Amount', 'Description']
    for col_idx, header in enumerate(headers1, 1):
        cell = ws1.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
    # Row 4 (openpyxl Row 5): Empty spacer row
    
    # Data rows start at Row 6
    row_idx1 = 6
    total_amount_sum = 0.0
    received_amount_sum = 0.0
    balance_amount_sum = 0.0
    
    for tx in transactions:
        tot_amt = safe_float(tx.get("total_amount"))
        rec_amt = safe_float(tx.get("paid_amount")) if tx.get("paid_amount") is not None else safe_float(tx.get("total_amount"))
        bal_amt = safe_float(tx.get("balance")) if tx.get("balance") is not None else 0.0
        
        ws1.cell(row=row_idx1, column=1, value=format_tx_date(tx)).alignment = Alignment(horizontal="center")
        ws1.cell(row=row_idx1, column=2, value=tx.get("name", "")).alignment = Alignment(horizontal="left")
        ws1.cell(row=row_idx1, column=3, value=tx.get("phone", "")).alignment = Alignment(horizontal="center")
        ws1.cell(row=row_idx1, column=4, value=tx.get("gstin", "")) # Party GSTIN
        ws1.cell(row=row_idx1, column=5, value=tx.get("order_no", "")) # Order No
        ws1.cell(row=row_idx1, column=6, value=get_invoice_number_formatted(tx)).alignment = Alignment(horizontal="center")
        ws1.cell(row=row_idx1, column=7, value="Sale").alignment = Alignment(horizontal="center")
        
        c8 = ws1.cell(row=row_idx1, column=8, value=f"{tot_amt:.2f}")
        c8.alignment = Alignment(horizontal="right")
        
        ws1.cell(row=row_idx1, column=9, value=tx.get("transaction_id", "")).alignment = Alignment(horizontal="center")
        
        c10 = ws1.cell(row=row_idx1, column=10, value=f"{rec_amt:.2f}")
        c10.alignment = Alignment(horizontal="right")
        
        c11 = ws1.cell(row=row_idx1, column=11, value=f"{bal_amt:.2f}")
        c11.alignment = Alignment(horizontal="right")
        
        ws1.cell(row=row_idx1, column=12, value=tx.get("description", "")).alignment = Alignment(horizontal="left")
        
        # Apply data font & border to all cells in the row
        for c in range(1, 13):
            cell = ws1.cell(row=row_idx1, column=c)
            cell.font = data_font
            cell.border = thin_border
            
        total_amount_sum += tot_amt
        received_amount_sum += rec_amt
        balance_amount_sum += bal_amt
        row_idx1 += 1
        
    # Spacer row
    for c in range(1, 13):
        ws1.cell(row=row_idx1, column=c).border = thin_border
    row_idx1 += 1
    
    # Totals Row
    ws1.cell(row=row_idx1, column=7, value="Total").alignment = Alignment(horizontal="center")
    ws1.cell(row=row_idx1, column=8, value=f"{total_amount_sum:.2f}").alignment = Alignment(horizontal="right")
    ws1.cell(row=row_idx1, column=10, value=f"{received_amount_sum:.2f}").alignment = Alignment(horizontal="right")
    ws1.cell(row=row_idx1, column=11, value=f"{balance_amount_sum:.2f}").alignment = Alignment(horizontal="right")
    
    for c in range(1, 13):
        cell = ws1.cell(row=row_idx1, column=c)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thin_border
        
    # ----------------------------------------------------
    # SHEET 2: Sale Items (Line-item details breakdown)
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="Sale Items")
    
    # Row 0 (openpyxl Row 1): Username metadata
    ws2.cell(row=1, column=1, value="Username").font = Font(name="Calibri", size=10, bold=True)
    ws2.cell(row=1, column=2, value=current_user["username"]).font = Font(name="Calibri", size=10)
    
    # Row 1 (openpyxl Row 2): Column Headers
    headers2 = ['Date', 'Party Name', 'Invoice No.', 'Item Name', 'Item code', 'HSN/SAC', 'Quantity', 'Unit', 'Price/Unit', 'Discount', 'GST', 'Amount']
    for col_idx, header in enumerate(headers2, 1):
        cell = ws2.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
    # Data rows start at Row 3
    row_idx2 = 3
    
    for tx in transactions:
        tx_date = format_tx_date(tx)
        party_name = tx.get("name", "")
        invoice_no = get_invoice_number_formatted(tx)
        
        items = tx.get("invoice_items", [])
        if not items:
            # Fallback to single dummy item
            p_name = tx.get("product", "")
            is_serv = False
            lower_p_name = p_name.lower()
            if any(x in lower_p_name for x in ["healing", "session", "class", "course", "reading", "appointment"]):
                is_serv = True
            unit = "Nos" if is_serv else "1"
            
            items = [{
                "name": p_name,
                "qty": 1,
                "price": safe_float(tx.get("amount")),
                "gst_rate": safe_float(tx.get("gst_rate")),
                "gst_amount": safe_float(tx.get("gst_total")),
                "total": safe_float(tx.get("total_amount")),
                "hsn": tx.get("hsn_code", "9983"),
                "unit": unit
            }]
            
        for item in items:
            item_name = item.get("name", "")
            hsn = item.get("hsn", item.get("hsn_code", ""))
            qty = safe_float(item.get("qty"), 1.0)
            price_unit = safe_float(item.get("price"))
            gst_rate = safe_float(item.get("gst_rate"))
            gst_amt = safe_float(item.get("gst_amount"))
            amount = safe_float(item.get("total"))
            
            # Determine unit
            item_unit = item.get("unit")
            if not item_unit:
                is_serv = service_map.get(item_name.lower().strip(), False)
                if not is_serv:
                    lower_item_name = item_name.lower()
                    if any(x in lower_item_name for x in ["healing", "session", "class", "course", "reading", "appointment"]):
                        is_serv = True
                item_unit = "Nos" if is_serv else "1"
                
            # Discount format: value(percentage%) e.g. 0.00(0.0%)
            disc_val = safe_float(item.get("discount_amount"))
            disc_pct = safe_float(item.get("discount_rate"))
            discount_str = f"{disc_val:.2f}({disc_pct:.1f}%)"
            
            # GST format: value(percentage%) e.g. 5.55(5.0%)
            gst_str = f"{gst_amt:.2f}({gst_rate:.1f}%)"
            
            ws2.cell(row=row_idx2, column=1, value=tx_date).alignment = Alignment(horizontal="center")
            ws2.cell(row=row_idx2, column=2, value=party_name).alignment = Alignment(horizontal="left")
            ws2.cell(row=row_idx2, column=3, value=invoice_no).alignment = Alignment(horizontal="center")
            ws2.cell(row=row_idx2, column=4, value=item_name).alignment = Alignment(horizontal="left")
            ws2.cell(row=row_idx2, column=5, value=item.get("code", "")) # Item code
            ws2.cell(row=row_idx2, column=6, value=hsn).alignment = Alignment(horizontal="center")
            ws2.cell(row=row_idx2, column=7, value=f"{qty:.1f}").alignment = Alignment(horizontal="right")
            ws2.cell(row=row_idx2, column=8, value=str(item_unit)).alignment = Alignment(horizontal="center")
            ws2.cell(row=row_idx2, column=9, value=f"{price_unit:.2f}").alignment = Alignment(horizontal="right")
            ws2.cell(row=row_idx2, column=10, value=discount_str).alignment = Alignment(horizontal="right")
            ws2.cell(row=row_idx2, column=11, value=gst_str).alignment = Alignment(horizontal="right")
            ws2.cell(row=row_idx2, column=12, value=f"{amount:.2f}").alignment = Alignment(horizontal="right")
            
            for c in range(1, 13):
                cell = ws2.cell(row=row_idx2, column=c)
                cell.font = data_font
                cell.border = thin_border
                
            row_idx2 += 1
            
    # Auto-fit column widths
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    # Reconstruct filename matching frontend
    filter_mode = payload.get("filter_mode", "date")
    if (filter_mode == 'date' or filter_mode == 'both') and start_date and end_date:
        def format_d(d_str):
            parts = d_str.split('-')
            return f"{parts[2]}-{parts[1]}-{parts[0]}"
        filename = f"Sale_Report_{format_d(start_date)}_to_{format_d(end_date)}.xlsx"
    else:
        today = datetime.now()
        filename = f"Sale_Report_{today.strftime('%d-%m-%Y')}.xlsx"

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Write to export_audit_logs collection
    from datetime import timezone
    item_count_calc = 0
    for tx in transactions:
        tx_items = tx.get("invoice_items", [])
        if not tx_items:
            item_count_calc += 1
        else:
            item_count_calc += len(tx_items)
            
    audit_log = {
        "exported_by": current_user.get("username", "admin"),
        "user_id": str(current_user.get("_id", "unknown")),
        "exported_at": datetime.now(timezone.utc),
        "filter_mode": filter_mode,
        "date_range": {
            "start": start_date or "",
            "end": end_date or ""
        },
        "selected_products": products,
        "transaction_count": len(transactions),
        "item_count": item_count_calc,
        "total_sales": round(total_amount_sum, 2),
        "filename": filename,
        "status": "SUCCESS"
    }
    await db.export_audit_logs.insert_one(audit_log)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.post("/export-preview")
async def export_preview(payload: dict, db=Depends(get_db), current_user=Depends(get_current_user)):
    """Generate live preview data of transactions matching the current filters."""
    from fastapi.responses import JSONResponse
    
    start_date = payload.get("start_date")  # "YYYY-MM-DD"
    end_date = payload.get("end_date")      # "YYYY-MM-DD"
    products = payload.get("products", [])   # list of product name strings
    
    # Defensive backend date validation
    if start_date and end_date:
        if start_date > end_date:
            return JSONResponse(
                status_code=400,
                content={
                    "success": False,
                    "message": "End date cannot be earlier than start date."
                }
            )
        
    query = {}
    
    # Date filter — parse date field (stored as DD/MM/YY string) via timestamp
    if start_date or end_date:
        date_filter = {}
        if start_date:
            try:
                date_filter["$gte"] = datetime.strptime(start_date, "%Y-%m-%d")
            except ValueError:
                pass
        if end_date:
            try:
                # End of day
                end_dt = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
                date_filter["$lte"] = end_dt
            except ValueError:
                pass
        if date_filter:
            query["timestamp"] = date_filter
    
    # Product filter — match any product in the items string
    if products and len(products) > 0:
        product_patterns = [{"product": {"$regex": re.escape(p), "$options": "i"}} for p in products]
        if "$or" in query:
            existing_or = query.pop("$or")
            query["$and"] = [{"$or": existing_or}, {"$or": product_patterns}]
        else:
            query["$or"] = product_patterns

    # Retrieve only the necessary fields for lightweight execution
    cursor = db.transactions.find(
        query, 
        {
            "total_amount": 1, 
            "paid_amount": 1, 
            "balance": 1, 
            "invoice_items": 1
        }
    )
    transactions = await cursor.to_list(length=100000)
    
    transaction_count = len(transactions)
    item_count = 0
    total_sales = 0.0
    received_amount = 0.0
    balance_amount = 0.0
    
    for tx in transactions:
        tot_amt = float(tx.get("total_amount", 0.0))
        rec_amt = float(tx.get("paid_amount", tx.get("total_amount", 0.0)) if tx.get("paid_amount") is not None else tx.get("total_amount", 0.0))
        bal_amt = float(tx.get("balance", 0.0) if tx.get("balance") is not None else 0.0)
        
        total_sales += tot_amt
        received_amount += rec_amt
        balance_amount += bal_amt
        
        items = tx.get("invoice_items", [])
        if not items:
            item_count += 1
        else:
            item_count += len(items)
            
    return {
        "success": True,
        "transaction_count": transaction_count,
        "item_count": item_count,
        "total_sales": round(total_sales, 2),
        "received_amount": round(received_amount, 2),
        "balance_amount": round(balance_amount, 2)
    }

@router.post("/export-suggestions")
async def export_suggestions(payload: dict, db=Depends(get_db), current_user=Depends(get_current_user)):
    """Generate smart recovery suggestions when export filters return 0 results."""
    from datetime import datetime, timedelta
    import calendar

    start_date = payload.get("start_date")
    end_date = payload.get("end_date")
    products = payload.get("products", [])
    filter_mode = payload.get("filter_mode", "date")

    # 1. Find nearest transaction date
    nearest_date_str = None
    start_dt = None
    end_dt = None

    if start_date:
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        except ValueError:
            pass
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
        except ValueError:
            pass

    # Query nearest transaction
    nearest_tx = None
    if start_dt and end_dt:
        # Search before start_dt
        prev_txs = await db.transactions.find({"timestamp": {"$lt": start_dt}}, {"timestamp": 1}).sort([("timestamp", -1)]).to_list(length=1)
        # Search after end_dt
        next_txs = await db.transactions.find({"timestamp": {"$gt": end_dt}}, {"timestamp": 1}).sort([("timestamp", 1)]).to_list(length=1)
        
        if prev_txs and next_txs:
            prev_diff = start_dt - prev_txs[0]["timestamp"]
            next_diff = next_txs[0]["timestamp"] - end_dt
            nearest_tx = prev_txs[0] if prev_diff < next_diff else next_txs[0]
        elif prev_txs:
            nearest_tx = prev_txs[0]
        elif next_txs:
            nearest_tx = next_txs[0]
            
    if not nearest_tx:
        # Fallback to absolute latest transaction
        latest_txs = await db.transactions.find({}, {"timestamp": 1}).sort([("timestamp", -1)]).to_list(length=1)
        if latest_txs:
            nearest_tx = latest_txs[0]

    if nearest_tx and "timestamp" in nearest_tx:
        nearest_date_str = nearest_tx["timestamp"].strftime("%Y-%m-%d")

    # 2. Recommended date range (month containing nearest date)
    rec_range = None
    if nearest_date_str:
        nearest_dt = datetime.strptime(nearest_date_str, "%Y-%m-%d")
        year = nearest_dt.year
        month = nearest_dt.month
        last_day = calendar.monthrange(year, month)[1]
        rec_range = {
            "start": f"{year}-{month:02d}-01",
            "end": f"{year}-{month:02d}-{last_day:02d}"
        }

    # 3. Available financial years
    first_txs = await db.transactions.find({}, {"timestamp": 1}).sort([("timestamp", 1)]).to_list(length=1)
    last_txs = await db.transactions.find({}, {"timestamp": 1}).sort([("timestamp", -1)]).to_list(length=1)
    
    financial_years = []
    if first_txs and last_txs:
        first_year = first_txs[0]["timestamp"].year
        last_year = last_txs[0]["timestamp"].year
        
        # Indian Financial Year starts April 1st
        # So check years from first_year - 1 to last_year + 1
        for y in range(first_year - 1, last_year + 2):
            # FY starts April 1st of y and ends March 31st of y + 1
            fy_label = f"FY {y - 2000:02d}-{y - 2000 + 1:02d}"
            # Check if there are transactions in this range
            fy_start = datetime(y, 4, 1)
            fy_end = datetime(y + 1, 3, 31, 23, 59, 59)
            tx_in_fy = await db.transactions.find({"timestamp": {"$gte": fy_start, "$lte": fy_end}}, {"_id": 1}).to_list(length=1)
            if tx_in_fy:
                financial_years.append(fy_label)

    # 4. Available products (top active or all)
    available_products = []
    products_cursor = db.products.find({}, {"name": 1})
    products_list = await products_cursor.to_list(length=100)
    available_products = [p["name"] for p in products_list if "name" in p]

    return {
        "success": True,
        "nearest_transaction_date": nearest_date_str,
        "available_financial_years": financial_years,
        "recommended_date_range": rec_range,
        "available_products": available_products
    }
