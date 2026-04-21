import openpyxl
import os

file_path = r"c:\Users\prane\Desktop\zeal_healing(web)\BILLING_WEB\billing-system\final_billing_20_users (1).xlsx"

if not os.path.exists(file_path):
    print("File not found")
else:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    products_found = set()
    for sheet in wb.worksheets:
        header = [str(c.value).lower() for c in next(sheet.iter_rows(min_row=1, max_row=1), [])]
        if "items" in header:
            idx = header.index("items")
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[idx]:
                    items = str(row[idx]).split(",")
                    for item in items:
                        name = item.strip()
                        if "(" in name:
                            name = name.split("(")[0].strip()
                        products_found.add(name)
        elif len(header) >= 6: # Old format fallback
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[5]:
                    products_found.add(str(row[5]).strip())
    
    print("Products in Excel:")
    for p in sorted(list(products_found)):
        print(f"- {p}")
